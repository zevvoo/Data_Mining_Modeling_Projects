######################################
## Microbiome Disease Prediction   ##
## Implemented by Jewoo Yoo         ##
## 2026-06-04                       ##
######################################

from pathlib import Path
import json
import os
import warnings

import joblib
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import ExtraTreesClassifier, RandomForestClassifier, VotingClassifier
from sklearn.feature_selection import SelectKBest, VarianceThreshold, f_classif
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold, cross_validate
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC


warnings.filterwarnings("ignore")

BASE_DIR = Path(__file__).resolve().parent
SOURCE_DIR = Path(os.environ.get("MICROBIOME_SOURCE_DIR", BASE_DIR / "input_data"))
DATA_DIR = BASE_DIR / "data"
MODEL_DIR = BASE_DIR / "model"
DATA_DIR.mkdir(parents=True, exist_ok=True)
MODEL_DIR.mkdir(parents=True, exist_ok=True)

TRAIN_PATH = SOURCE_DIR / "Q2_train.tsv"
TEST_PATH = SOURCE_DIR / "Q2_test.tsv"
ANSWER_TEMPLATE = SOURCE_DIR / "Q2_AnswerSheet.xlsx"

PREDICTION_CSV = DATA_DIR / "Q2_test_predictions.csv"
METRICS_CSV = DATA_DIR / "model_metrics_cv.csv"
ANSWER_OUTPUT = BASE_DIR / "Q2_AnswerSheet_filled.xlsx"
MODEL_OUTPUT = MODEL_DIR / "microbiome_disease_voting_model.joblib"
SUMMARY_OUTPUT = DATA_DIR / "modeling_summary.json"


def make_extra_trees(seed=2026, n_estimators=1000):
    return Pipeline(
        steps=[
            ("var", VarianceThreshold(1e-12)),
            (
                "clf",
                ExtraTreesClassifier(
                    n_estimators=n_estimators,
                    random_state=seed,
                    class_weight="balanced",
                    n_jobs=-1,
                    max_features="sqrt",
                ),
            ),
        ]
    )


def make_random_forest(seed=2026):
    return Pipeline(
        steps=[
            ("var", VarianceThreshold(1e-12)),
            (
                "clf",
                RandomForestClassifier(
                    n_estimators=800,
                    random_state=seed,
                    class_weight="balanced",
                    n_jobs=-1,
                    max_features="sqrt",
                ),
            ),
        ]
    )


def make_logistic():
    return Pipeline(
        steps=[
            ("var", VarianceThreshold(1e-12)),
            ("scaler", StandardScaler()),
            ("select", SelectKBest(f_classif, k=500)),
            (
                "clf",
                LogisticRegression(
                    max_iter=3000,
                    C=1.0,
                    class_weight="balanced",
                    solver="lbfgs",
                ),
            ),
        ]
    )


def make_svc():
    return Pipeline(
        steps=[
            ("var", VarianceThreshold(1e-12)),
            ("scaler", StandardScaler()),
            ("select", SelectKBest(f_classif, k=500)),
            (
                "clf",
                SVC(
                    C=3,
                    gamma="scale",
                    class_weight="balanced",
                    probability=True,
                ),
            ),
        ]
    )


def make_voting_model():
    return VotingClassifier(
        estimators=[
            ("extra_trees", make_extra_trees(seed=2026, n_estimators=1000)),
            ("logistic", make_logistic()),
        ],
        voting="soft",
        weights=[4, 1],
        n_jobs=1,
    )


def evaluate_models(X, y):
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=2026)
    models = {
        "ExtraTrees": make_extra_trees(seed=2026, n_estimators=1000),
        "RandomForest": make_random_forest(seed=2026),
        "LogisticRegression": make_logistic(),
        "SVC_RBF": make_svc(),
        "SoftVoting_ET4_LR1": make_voting_model(),
    }

    rows = []
    for name, model in models.items():
        result = cross_validate(
            model,
            X,
            y,
            cv=cv,
            scoring={"accuracy": "accuracy", "balanced_accuracy": "balanced_accuracy"},
            n_jobs=1,
        )
        rows.append(
            {
                "model": name,
                "accuracy_mean": result["test_accuracy"].mean(),
                "accuracy_sd": result["test_accuracy"].std(),
                "balanced_accuracy_mean": result["test_balanced_accuracy"].mean(),
                "balanced_accuracy_sd": result["test_balanced_accuracy"].std(),
            }
        )
    metrics = pd.DataFrame(rows).sort_values("accuracy_mean", ascending=False)
    metrics.to_csv(METRICS_CSV, index=False, encoding="utf-8-sig")
    return metrics


def fill_answer_sheet(test_ids, predictions):
    wb = load_workbook(ANSWER_TEMPLATE)
    ws = wb.active

    header = [ws.cell(row=1, column=1).value, ws.cell(row=1, column=2).value]
    if header != ["ID", "Prediction"]:
        raise ValueError(f"Unexpected answer sheet header: {header}")

    id_to_prediction = dict(zip(test_ids, predictions))
    for row in range(2, ws.max_row + 1):
        sample_id = ws.cell(row=row, column=1).value
        ws.cell(row=row, column=2).value = id_to_prediction[sample_id]

    wb.save(ANSWER_OUTPUT)


def main():
    train = pd.read_csv(TRAIN_PATH, sep="\t")
    test = pd.read_csv(TEST_PATH, sep="\t")

    feature_cols = [c for c in train.columns if c != "Disease"]
    if list(test.columns[1:]) != feature_cols:
        raise ValueError("Train and test feature columns do not match.")

    X = train[feature_cols]
    y = train["Disease"]
    X_test = test[feature_cols]

    metrics = evaluate_models(X, y)
    best_model_name = metrics.iloc[0]["model"]
    final_model = make_voting_model()
    final_model.fit(X, y)

    predictions = final_model.predict(X_test)
    probabilities = final_model.predict_proba(X_test)
    max_probability = probabilities.max(axis=1)

    prediction_df = pd.DataFrame(
        {
            "ID": test["ID"],
            "Prediction": predictions,
            "max_probability": max_probability,
        }
    )
    prediction_df.to_csv(PREDICTION_CSV, index=False, encoding="utf-8-sig")
    fill_answer_sheet(test["ID"], predictions)

    bundle = {
        "model": final_model,
        "feature_cols": feature_cols,
        "classes": list(final_model.classes_),
    }
    joblib.dump(bundle, MODEL_OUTPUT)

    summary = {
        "train_rows": int(train.shape[0]),
        "test_rows": int(test.shape[0]),
        "feature_count": len(feature_cols),
        "class_count": int(y.nunique()),
        "best_cv_model": best_model_name,
        "selected_final_model": "SoftVoting_ET4_LR1",
        "selected_cv_accuracy": float(metrics[metrics["model"] == "SoftVoting_ET4_LR1"]["accuracy_mean"].iloc[0]),
        "selected_cv_balanced_accuracy": float(
            metrics[metrics["model"] == "SoftVoting_ET4_LR1"]["balanced_accuracy_mean"].iloc[0]
        ),
        "answer_sheet": str(ANSWER_OUTPUT),
        "prediction_csv": str(PREDICTION_CSV),
        "model_file": str(MODEL_OUTPUT),
    }
    SUMMARY_OUTPUT.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    print("Microbiome disease prediction modeling summary")
    print(f"Train rows: {summary['train_rows']}")
    print(f"Test rows: {summary['test_rows']}")
    print(f"Features: {summary['feature_count']}")
    print(f"Classes: {summary['class_count']}")
    print(f"Selected model: {summary['selected_final_model']}")
    print(f"5-fold CV accuracy: {summary['selected_cv_accuracy']:.4f}")
    print(f"5-fold CV balanced accuracy: {summary['selected_cv_balanced_accuracy']:.4f}")
    print(f"Answer sheet written: {ANSWER_OUTPUT}")


if __name__ == "__main__":
    main()
